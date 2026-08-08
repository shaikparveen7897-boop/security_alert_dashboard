"""
excel_report.py
----------------
Generates a clean, presentation-ready .xlsx report of login events for
download from the dashboard. Built with openpyxl so it works fully offline
(no external services), matching the rest of the project.
"""

from io import BytesIO
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2530", end_color="1F2530", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10.5)
FAILED_FILL = PatternFill(start_color="FCE8E7", end_color="FCE8E7", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="E7F6F3", end_color="E7F6F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9DEE3"),
    right=Side(style="thin", color="D9DEE3"),
    top=Side(style="thin", color="D9DEE3"),
    bottom=Side(style="thin", color="D9DEE3"),
)

COLUMNS = [
    ("Timestamp (UTC)", 22),
    ("Username", 18),
    ("IP Address", 16),
    ("Location", 20),
    ("Status", 12),
    ("Reason", 20),
    ("Flagged Suspicious", 18),
]


def build_report(events, suspicious_ips, generated_by="Sentry Dashboard", start_iso=None, end_iso=None):
    """
    Builds an in-memory .xlsx workbook from the given events and returns it
    as BytesIO, ready to be sent as a Flask file download.
    """
    suspicious_set = {s["ip"] for s in suspicious_ips}

    wb = Workbook()
    ws = wb.active
    ws.title = "Login Events"

    # --- Title / metadata block ---
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sentry — Security Alert Report"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)

    range_label = "All available data"
    if start_iso or end_iso:
        range_label = f"{start_iso or 'earliest'}  to  {end_iso or 'latest'}"

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Report range: {range_label}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')} by {generated_by}"
    ws["A3"].font = Font(name="Arial", size=10, italic=True, color="555555")

    header_row = 5

    # --- Header row ---
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    # --- Data rows ---
    sorted_events = sorted(events, key=lambda e: e["timestamp"], reverse=True)
    row = header_row + 1
    for e in sorted_events:
        is_flagged = e["ip"] in suspicious_set and e["status"] == "failed"
        values = [
            e["timestamp"].replace("T", " ").split("+")[0],
            e["username"],
            e["ip"],
            e["country"],
            e["status"].upper(),
            e["reason"].replace("_", " ").title(),
            "YES" if is_flagged else "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
            if e["status"] == "failed":
                cell.fill = FAILED_FILL
            else:
                cell.fill = SUCCESS_FILL
        row += 1

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Detection Summary"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True)

    total = len(events)
    failed = sum(1 for e in events if e["status"] == "failed")
    success = total - failed

    summary_rows = [
        ("Report range", range_label),
        ("Total login events", total),
        ("Failed attempts", failed),
        ("Successful logins", success),
        ("Flagged suspicious IPs", len(suspicious_ips)),
        ("Detection rule", "5+ failed attempts from one IP within a 15-minute window"),
    ]
    for i, (label, value) in enumerate(summary_rows, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True)
        ws2.cell(row=i, column=2, value=value).font = Font(name="Arial")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 50

    if suspicious_ips:
        start_r = len(summary_rows) + 5
        ws2.cell(row=start_r, column=1, value="Flagged IPs").font = Font(name="Arial", size=12, bold=True)
        headers = ["IP Address", "Location", "Failed Attempts", "Targeted Users", "Risk Level"]
        for c, h in enumerate(headers, start=1):
            cell = ws2.cell(row=start_r + 1, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for i, s in enumerate(suspicious_ips, start=start_r + 2):
            ws2.cell(row=i, column=1, value=s["ip"])
            ws2.cell(row=i, column=2, value=s["country"])
            ws2.cell(row=i, column=3, value=s["failed_count"])
            ws2.cell(row=i, column=4, value=", ".join(s["usernames_targeted"]))
            ws2.cell(row=i, column=5, value=s["risk_level"].upper())

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
